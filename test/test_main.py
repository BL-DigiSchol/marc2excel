# -*- coding: utf8 -*-

import os
import openpyxl
import pymarc
import filecmp
import tempfile
import marc2excel
from marc2excel.main import Converter
from pytest_mock import mocker

converter = marc2excel.Converter()


class TestConverter():

    def test_excel2marc_convenience_method(self, tmp_mrc, tmp_xlsx, mocker):
        mocker.patch('marc2excel.main.Converter.excel2marc')
        marc2excel.excel2marc(tmp_xlsx, tmp_mrc)
        Converter.excel2marc.assert_called_once_with(tmp_xlsx, tmp_mrc,
                                                     force_utf8=False, sheet=0)

    def test_marc2excel_convenience_method(self, tmp_mrc, tmp_xlsx, mocker):
        mocker.patch('marc2excel.main.Converter.marc2excel')
        marc2excel.marc2excel(tmp_mrc, tmp_xlsx)
        Converter.marc2excel.assert_called_once_with(tmp_mrc, tmp_xlsx,
                                                     force_utf8=False)

    def test_marc2excel_writes_all_records(self, marc_path, tmp_xlsx):
        converter.marc2excel(marc_path, tmp_xlsx)
        wb = openpyxl.load_workbook(tmp_xlsx)
        ws = wb.get_active_sheet()
        assert ws.max_row == 21
        assert ws.max_column == 80

    def test_marc2excel_and_excel2marc_compatible(self, marc_path, tmp_xlsx,
                                                  tmp_mrc):
        converter.marc2excel(marc_path, tmp_xlsx)
        converter.excel2marc(tmp_xlsx, tmp_mrc.name)
        assert list(open(marc_path, 'r')) == list(open(tmp_mrc.name, 'r'))

